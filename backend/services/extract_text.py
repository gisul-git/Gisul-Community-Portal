

import pdfplumber
import io
import logging
import os
import mimetypes
import re
import tempfile
import subprocess
from pathlib import Path

OCR_AVAILABLE = False
TESSERACT_INSTALLED = False

try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
    print("✅ OCR libraries (pytesseract, Pillow) are installed")
    
    try:
        pytesseract.get_tesseract_version()
        TESSERACT_INSTALLED = True
        print("✅ Tesseract OCR engine is installed and available")
    except Exception as tesseract_check_err:
        print(f"⚠️  Tesseract OCR engine not found: {tesseract_check_err}")
        print("   Please install Tesseract OCR:")
        print("   - Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki")
        print("   - Linux: sudo apt-get install tesseract-ocr")
        print("   - macOS: brew install tesseract")
        TESSERACT_INSTALLED = False
except ImportError as import_err:
    OCR_AVAILABLE = False
    print(f"⚠️  OCR libraries (pytesseract, Pillow) not installed: {import_err}")
    print("   Install with: pip install pytesseract Pillow pdf2image")

try:
    from pdf2image import convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False
    print("⚠️  pdf2image not installed. Scanned PDF OCR will be limited.")
    print("   Install with: pip install pdf2image")
    print("   Note: On Linux, you may also need: sudo apt-get install poppler-utils")

EXCEL_AVAILABLE = False
try:
    import openpyxl
    EXCEL_AVAILABLE = True
    print("✅ Excel support (openpyxl) is available")
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel file support will be limited.")
    print("   Install with: pip install openpyxl")

# For older .xls files
XLS_AVAILABLE = False
try:
    import xlrd
    XLS_AVAILABLE = True
    print("✅ Legacy Excel support (xlrd) is available")
except ImportError:
    XLS_AVAILABLE = False
    print("⚠️  xlrd not installed. Legacy .xls file support will be limited.")
    print("   Install with: pip install xlrd")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def clean_text(text: str) -> str:
    
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'\n\s*\n', '\n', text)
    return text.strip()

def convert_doc_to_pdf(file_path: str) -> str:
    
    try:
        out_dir = tempfile.mkdtemp()
        base_name = Path(file_path).stem
        pdf_target = Path(out_dir) / f"{base_name}.pdf"

        cmd_primary = [
            "soffice", "--headless",
            "--convert-to", "pdf",
            "--outdir", out_dir,
            file_path
        ]
        logger.warning(f"🧾 Converting Word file to PDF: {file_path}")
        res1 = subprocess.run(cmd_primary, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60)

        if pdf_target.exists():
            logger.warning(f"✅ Converted to PDF: {pdf_target}")
            return str(pdf_target)

        logger.warning("⚠️ Primary conversion failed. Retrying with 'writer_pdf_Export' filter...")
        cmd_retry = [
            "soffice", "--headless",
            "--convert-to", "pdf:writer_pdf_Export",
            "--outdir", out_dir,
            file_path
        ]
        res2 = subprocess.run(cmd_retry, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60)

        if pdf_target.exists():
            logger.warning(f"✅ Retry successful using 'writer_pdf_Export': {pdf_target}")
            return str(pdf_target)

        logger.warning(f"❌ Failed to convert {file_path} to PDF")
        logger.debug(f"stderr: {res1.stderr.decode(errors='ignore')} | {res2.stderr.decode(errors='ignore')}")
        return ""

    except subprocess.TimeoutExpired:
        logger.warning(f"⏱️ Conversion timed out for {file_path}")
        return ""
    except Exception as e:
        logger.warning(f"❌ Conversion error for {file_path}: {e}")
        return ""

def try_catdoc(file_path: str) -> str:
    
    try:
        result = subprocess.run(
            ["catdoc", "-d", "utf-8", file_path],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30
        )
        if result.returncode == 0 and result.stdout:
            text = result.stdout.decode("utf-8", errors="ignore")
            logger.warning(f"✅ Extracted text via Catdoc (length: {len(text)})")
            return text
        else:
            logger.warning(f"⚠️ Catdoc failed: {result.stderr.decode(errors='ignore')}")
            return ""
    except FileNotFoundError:
        logger.warning("⚠️ Catdoc not found in PATH.")
        return ""
    except Exception as e:
        logger.warning(f"❌ Catdoc extraction error: {e}")
        return ""

def extract_text_from_pdf(pdf_path: str) -> str:
    
    try:
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text.strip()
    except Exception as e:
        logger.warning(f"⚠️ PDF text extraction failed: {e}")
        return ""

def ocr_pdf_file(pdf_path: str) -> str:
    
    if not OCR_AVAILABLE or not PDF2IMAGE_AVAILABLE:
        logger.warning("⚠️ OCR not available for PDF processing")
        return ""
    
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(pdf_path, dpi=300)
        logger.info(f"📷 Converted PDF to {len(images)} image(s) for OCR")
        
        ocr_text = ""
        for i, image in enumerate(images):
            logger.info(f"🔍 Running OCR on page {i+1}/{len(images)}...")
            page_text = pytesseract.image_to_string(image, lang='eng')
            ocr_text += page_text.strip() + "\n"
        
        return ocr_text.strip()
    except Exception as e:
        logger.warning(f"⚠️ OCR extraction from PDF failed: {e}")
        return ""

def extract_text_with_ocr(image_bytes: bytes) -> str:
    
    if not OCR_AVAILABLE:
        error_msg = "OCR not available - pytesseract or Pillow not installed. Please install: pip install pytesseract Pillow"
        logger.error(error_msg)
        print(f"❌ {error_msg}")
        raise ValueError(error_msg)
    
    if not TESSERACT_INSTALLED:
        error_msg = "Tesseract OCR engine not found. Please install Tesseract OCR on your system."
        logger.error(error_msg)
        print(f"❌ {error_msg}")
        raise RuntimeError(error_msg)
    
    try:
        try:
            version = pytesseract.get_tesseract_version()
            logger.info(f"Using Tesseract version: {version}")
        except Exception as tesseract_err:
            error_msg = f"Tesseract OCR engine not accessible. Error: {str(tesseract_err)}. Please ensure Tesseract is installed and in your PATH."
            logger.error(error_msg)
            print(f"❌ {error_msg}")
            raise RuntimeError(error_msg)
        
        image = Image.open(io.BytesIO(image_bytes))
        
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        logger.info(f"🔍 Running OCR on image (size: {image.size}, mode: {image.mode})...")
        
        text = pytesseract.image_to_string(image, lang='eng')
        
        text = text.strip()
        text = ' '.join(text.split())
        
        if not text:
            logger.warning("OCR returned empty text - image may not contain readable text or may be too low quality")
            raise ValueError("OCR could not extract any text from the image. Please ensure the image is clear, readable, and contains text.")
        
        logger.info(f"✅ OCR extracted {len(text)} characters from image")
        return text
    except (ValueError, RuntimeError):
        raise
    except Exception as e:
        error_msg = f"OCR extraction failed: {str(e)}. Please ensure Tesseract OCR is installed and the image is readable."
        logger.error(error_msg)
        import traceback
        logger.error(traceback.format_exc())
        raise RuntimeError(error_msg)

def extract_text_from_bytes(filename: str, file_bytes: bytes) -> str:
    
    try:
        file_lower = filename.lower()
        
        if file_lower.endswith(('.jpg', '.jpeg', '.png')):
            logger.info(f"📷 Detected image file: {filename}, using OCR")
            if not OCR_AVAILABLE:
                raise ValueError("OCR not available. Please install pytesseract and Pillow: pip install pytesseract Pillow")
            try:
                return extract_text_with_ocr(file_bytes)
            except (ValueError, RuntimeError) as ocr_err:
                logger.error(f"OCR failed for {filename}: {ocr_err}")
                raise
        
        elif file_lower.endswith('.pdf'):
            logger.info(f"📄 Detected PDF file: {filename}")
            
            try:
                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    text = ""
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                    
                    if text.strip() and len(text.strip()) > 50:
                        logger.info(f"✅ Extracted {len(text)} characters using standard PDF extraction")
                        return text.strip()
                    else:
                        logger.info(f"⚠️  PDF appears to be scanned/image-based (extracted only {len(text)} chars), trying OCR...")
            except Exception as pdf_err:
                logger.warning(f"Standard PDF extraction failed: {pdf_err}, trying OCR...")
            
            if OCR_AVAILABLE and PDF2IMAGE_AVAILABLE:
                try:
                    images = convert_from_bytes(file_bytes, dpi=300)
                    logger.info(f"📷 Converted PDF to {len(images)} image(s) for OCR")
                    
                    ocr_text = ""
                    for i, image in enumerate(images):
                        logger.info(f"🔍 Running OCR on page {i+1}/{len(images)}...")
                        page_text = pytesseract.image_to_string(image, lang='eng')
                        ocr_text += page_text.strip() + "\n"
                    
                    if ocr_text.strip():
                        logger.info(f"✅ OCR extracted {len(ocr_text)} characters from scanned PDF")
                        return ocr_text.strip()
                    else:
                        logger.warning("OCR returned empty text")
                except Exception as ocr_err:
                    logger.error(f"OCR extraction from PDF failed: {ocr_err}")
            else:
                logger.warning("OCR not available for scanned PDF processing")
            
            return text.strip() if 'text' in locals() else ""
        
        elif file_lower.endswith(('.doc', '.docx')):
            logger.warning(f"📝 Detected Word resume: {filename}")
            
            temp_doc_path = None
            try:
                if not file_bytes:
                    logger.error("⚠️ No file_bytes provided for Word document")
                    return ""
                
                suffix = '.doc' if file_lower.endswith('.doc') else '.docx'
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
                    temp_file.write(file_bytes)
                    temp_doc_path = temp_file.name
                
                pdf_path = convert_doc_to_pdf(temp_doc_path)

                if not pdf_path:
                    logger.warning(f"❌ Could not convert {filename} to PDF. Skipping.")
                    return ""

                text = extract_text_from_pdf(pdf_path)
                if not text or len(text) < 50:
                    logger.warning("⚠️ Extracted text is too short or empty. Trying OCR fallback...")
                    text = ocr_pdf_file(pdf_path)

                try:
                    os.remove(pdf_path)
                    pdf_dir = Path(pdf_path).parent
                    if pdf_dir.exists():
                        try:
                            os.rmdir(pdf_dir)
                        except:
                            pass
                except Exception:
                    pass

                if text:
                    logger.warning(f"✅ Successfully extracted text from converted PDF ({len(text)} chars)")
                    return clean_text(text)
                else:
                    logger.warning(f"⚠️ No text found even after OCR fallback for {filename}")
                    return ""
                    
            finally:
                if temp_doc_path and os.path.exists(temp_doc_path):
                    try:
                        os.unlink(temp_doc_path)
                    except Exception:
                        pass
        
        elif file_lower.endswith(('.xlsx', '.xls')):
            logger.info(f"📊 Detected Excel file: {filename}")
            
            if file_lower.endswith('.xlsx'):
                # Handle .xlsx files using openpyxl
                if not EXCEL_AVAILABLE:
                    raise ValueError("Excel support not available. Please install openpyxl: pip install openpyxl")
                
                try:
                    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                    text_parts = []
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        text_parts.append(f"Sheet: {sheet_name}\n")
                        
                        for row in sheet.iter_rows(values_only=True):
                            row_text = []
                            for cell in row:
                                if cell is not None:
                                    cell_str = str(cell).strip()
                                    if cell_str:
                                        row_text.append(cell_str)
                            
                            if row_text:
                                text_parts.append(" | ".join(row_text) + "\n")
                        
                        text_parts.append("\n")
                    
                    workbook.close()
                    text = "".join(text_parts).strip()
                    
                    if text:
                        logger.info(f"✅ Extracted {len(text)} characters from Excel file")
                        return clean_text(text)
                    else:
                        logger.warning("⚠️ Excel file appears to be empty")
                        return ""
                        
                except Exception as excel_err:
                    logger.error(f"❌ Excel extraction failed: {excel_err}")
                    raise ValueError(f"Failed to extract text from Excel file: {excel_err}")
            
            else:
                # Handle legacy .xls files using xlrd
                if not XLS_AVAILABLE:
                    raise ValueError("Legacy Excel support not available. Please install xlrd: pip install xlrd")
                
                try:
                    workbook = xlrd.open_workbook(file_contents=file_bytes)
                    text_parts = []
                    
                    for sheet in workbook.sheets():
                        text_parts.append(f"Sheet: {sheet.name}\n")
                        
                        for row_idx in range(sheet.nrows):
                            row_values = []
                            for col_idx in range(sheet.ncols):
                                cell = sheet.cell(row_idx, col_idx)
                                cell_value = cell.value
                                
                                if cell.ctype == xlrd.XL_CELL_DATE:
                                    # Handle date cells
                                    try:
                                        date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                                        cell_value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                                    except:
                                        cell_value = str(cell_value)
                                elif cell_value is not None:
                                    cell_value = str(cell_value).strip()
                                
                                if cell_value:
                                    row_values.append(cell_value)
                            
                            if row_values:
                                text_parts.append(" | ".join(row_values) + "\n")
                        
                        text_parts.append("\n")
                    
                    text = "".join(text_parts).strip()
                    
                    if text:
                        logger.info(f"✅ Extracted {len(text)} characters from legacy Excel file")
                        return clean_text(text)
                    else:
                        logger.warning("⚠️ Excel file appears to be empty")
                        return ""
                        
                except Exception as xls_err:
                    logger.error(f"❌ Legacy Excel extraction failed: {xls_err}")
                    raise ValueError(f"Failed to extract text from Excel file: {xls_err}")
        
        else:
            logger.info(f"📄 Attempting to decode as plain text: {filename}")
            return file_bytes.decode('utf-8', errors='ignore').strip()
            
    except Exception as e:
        logger.error(f"❌ Error extracting text from {filename}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return ""

